from openpyxl import load_workbook

def workbook_data_to_dict(workbook_path):
    wb = load_workbook(workbook_path)
    ws = wb.active

    max_rows = ws.max_row

    data = {}
    for row in range(2, max_rows + 1):
        data[ws[f"A{row}"].value] = ws[f"B{row}"].value

    return data
